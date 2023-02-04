import pywinauto
import time
import os
import pyautogui
import pandas
import openpyxl
from openpyxl import workbook
from openpyxl import load_workbook

wb = load_workbook("C:\\Automation\\TestCaseSelector_VrniAutomation.xlsx")
sh = wb.active
sh = wb['TestFunctions']
sh2 = wb['TestCase']

#s1 = sh.cell(2,2)
s2 = sh.cell(3,2)
s3 = sh.cell(4,2)
s4 = sh.cell(5,2)
s5 = sh.cell(6,2)
p1 = sh.cell(2,5)
p2 = sh.cell(3,5)
p3 = sh.cell(4,5)
p4 = sh.cell(5,5)
p5 = sh.cell(6,5)

q1 = sh2.cell(2,3)
q2 = sh2.cell(3,3)
q3 = sh2.cell(4,3)


f1 = sh.cell(2,1)
f2 = sh.cell(3,1)
f3 = sh.cell(4,1)
f4 = sh.cell(5,1)
f5 = sh.cell(6,1)

J1 = f1.value
J2  = f2.value
J3 = f3.value
J4 = f4.value
J5 = f5.value

#A = s1.value
B= s2.value
C = s3.value
D = s4.value
E = s5.value

G = p1.value
H = p2.value
I = p3.value
J = p4.value
K= p5.value

L = q1.value
M = q2.value
N = q3.value

wrkbk = load_workbook("C:\\Automation\\TestCaseInputFile.xlsx")
wb1 = load_workbook("C:\\Automation\\TestCaseInputFile2.xlsx")
wb2 = load_workbook("C:\\Automation\\TestCaseInputFile3.xlsx")

sh1 = wrkbk.active
sh3= wb1.active
sh4= wb2.active
c1= sh1.cell(row=2,column=7)
c2= sh1.cell(row=2,column=8)
c3= sh1.cell(row=2,column=9)
c4= sh1.cell(row=2,column=10)
c5= sh1.cell(row=2,column=11)
c6= sh1.cell(row=2,column=12)
c7= sh1.cell(row=2,column=13)
c8= sh1.cell(row=2,column=14)
c9= sh1.cell(row=2,column=15)
c10 = sh3.cell(row=2,column=11)
c11 = sh3.cell(row=2,column=12)
c12 = sh3.cell(row=2,column=13)
c13 = sh4.cell(row=2,column=14)
c14 = sh4.cell(row=2,column=15)
c15 = sh4.cell(row=2,column=12)
c16 = sh3.cell(row=2,column=14)
c17 = sh4.cell(row=2,column=16)
c18 = sh1.cell(row=2,column=16)

ConfigMgrIPAddress = c1.value
OVD =  c2.value
NWDR = c3.value
DEV =  c4.value
MOD1 =  c5.value
PIINJ = c6.value
NWPNT = c7.value
PARA = c8.value
OUT = c9.value
MOD3 = c10.value
NWPNT1 = c11.value
OUT1 = c12.value
NWPNT2 = c13.value
OUT2 = c14.value
ENA = c15.value
ANAPNT = c16.value
PAKDPNT = c17.value
DIGPNT = c18.value

app = pywinauto.application.Application(backend = "uia")
app1 = pywinauto.application.Application(backend = "uia")

import pandas as pd
import numpy as np
#df = pd.read_excel("C:\\Automation\\TestCaseSelector_VrniAutomation.xlsx",sheet_name='TestFunctions', usecols="A")
from openpyxl import Workbook, load_workbook
wb = load_workbook("C:\\Automation\\TestCaseSelector_VrniAutomation.xlsx")
source = wb["TestFunctions"]
source1 = wb["TestCase"]
#for cell in source['B']:
#  print(cell.value)


from package import vrni2
from package import vrni3
from package import vrni4

h1 = vrni2.VRNIFirst()
w1 = vrni3.VRNISecond()
w2 = vrni4.VRNIThird()

z= 0
row = 2
col = 2
wb = load_workbook("C:\\Automation\\TestCaseSelector_VrniAutomation.xlsx")
sh = wb.active
sh = wb['TestFunctions']
U1 = sh.cell(row,col)
row3 = 2
col3 = 6
U5 = sh.cell(row3,col3)
print(U5.value)

row1 = 2
col1 = 5
col3 = 3
U2 = sh.cell(row1,col1)

row2 = 2
col2 = 2

"""

for cell in source['F']
     if cell.value == "Continue"
"""
     





"""


for cell in source1['B']:
      if cell.value != None:
            U3 = sh2.cell(row2,col2)
            U4 = sh2.cell(row2,col3)
            print(U3.value)
            print(U4.value)
            row2 += 1
            #col2 += 1
            if U4.value == "Y":
                for cell in source['A']:
                      if cell.value == U3.value:
                            for cell in source['B']:
                                  U1 = sh.cell(row,col)
                                  U2 = sh.cell(row1,col1)
                                  #print(U1.value)
                                  #print(U2.value)
                                  row += 1
                                  row1 += 1
                                  if U1.value != None and U2.value == "Y" :
                                       if U1.value == "LaunchVRNI":
                                             h2 = h1.LaunchVRNI(app)
                                             print('VRNI has been launched succesfully')  
                                       elif U1.value == "ConnectVRNI":
                                             h3 = h1.ConnectVRNI(app,ConfigMgrIPAddress)
                                             print("ConnectVRNI keyword has been executed succesfully")
                                       elif U1.value == "AddVRNIDrop":
                                             h4 = h1.AddVRNIDrop(app,OVD,NWDR)
                                             print("AddVRNIDrop keyword has been executed successfully")
                                       elif U1.value == "DigitalValueInject":
                                             h5 = h1.DigitalValueInject(app,DEV,MOD1,PIINJ)
                                             print("DigitalValueInject keyword has been executed successfully")
                                       elif U1.value == "LaunchPIReadDigitalValue":
                                             h6 = h1.LaunchPIReadDigitalValue(app,app1,NWPNT,DIGPNT,PARA,OUT)
                                             print("LaunchPIReadDigitalValue keyword has been executed successfully")
                                       elif U1.value == "KillVRNIConfigMgr":
                                              h10 = h1.KillVRNIConfigMgr()
                                              print("VRNI ConfigMgr  window has been closed")
                                       elif U1.value == "KillPointInfo":
                                              h11 = h1.KillPointInfo()
                                              print("Point Info window has been closed now")
                                       elif U1.value == "AnalogValueInject":
                                             h7 = w1.AnalogValueInject(app,DEV,MOD3)
                                             print("AnalogValueInject keyword has been executed successfully")
                                       elif U1.value == "LaunchPIReadAnalogValue":
                                             h8 = w1.LaunchPIReadAnalogValue(app,app1,NWPNT1, ANAPNT,OUT1)
                                             print("LaunchPIReadAnalogValue keyword has been executed successfully")
                                       elif U1.value == "KillVRNIConfigMgr":
                                             h12 = w1.KillVRNIConfigMgr()
                                             print("VRNI ConfigMgr  window has been closed")
                                       elif U1.value == "KillPointInfo":
                                              h13 = w1.KillPointInfo()
                                              print("Point Info window has been closed now")
                                       elif U1.value == "PackedValueInject":
                                             h9 = w2.PackedValueInject(app,DEV,MOD1,ENA,PIINJ)
                                             print("PackedValueInject keyword has been executed successfully")
                                       elif U1.value == "LaunchPIReadPackedValue":
                                             h9 = w2.LaunchPIReadPackedValue(app,app1,NWPNT2, PAKDPNT, OUT2)
                                             print("LaunchPIReadPackedValue keyword has been executed successfully")
                                       elif U1.value == "KillVRNIConfigMgr":
                                             h13 = w2.KillVRNIConfigMgr()
                                             print("VRNI ConfigMgr  window has been closed")
                                       elif U1.value == "KillPointInfo":
                                             h14 = w2.KillPointInfo()
                                             print("Point Info window has been closed now")
                                             #break
                                       #break
                                  #break
                                         
                      
                  
"""         
                
                  
       
                              
                                    
                                    
                              
                                    
                              
                              
                                    
                              
                              
                        
                  
                  
      
                        

                           
                                  


                               
                               
                           
                               

                           
               



                

     
